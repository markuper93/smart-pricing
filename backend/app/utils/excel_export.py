import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

def generate_excel(report_data: List[Dict[str, Any]], title: str, avg_change: float, filename: str) -> str:
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31].replace(':', '-').replace('\\', '/').replace('?', '').replace('*', '').replace('[', '').replace(']', '')  # Excel sheet name limit
    
    # RTL sheet
    ws.sheet_view.rightToLeft = True
    
    # Header style
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    if not report_data:
        ws.cell(row=1, column=1, value="אין נתונים להצגה")
        wb.save(filepath)
        return filepath
    
    # Write headers
    headers = list(report_data[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Write data
    for row_idx, row_data in enumerate(report_data, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)
            # Try to set numeric format
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"
    
    # Auto-fit column widths
    for col_idx in range(1, len(headers)+1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(report_data)+2))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)
    
    # Average row
    avg_row = len(report_data) + 3
    ws.cell(row=avg_row, column=1, value=f"אחוז שינוי ממוצע לכלל הרכבים (מחירון ללא עליה לכביש): {avg_change:.2f}%")
    ws.cell(row=avg_row, column=1).font = Font(name="Arial", bold=True, size=11)
    ws.merge_cells(start_row=avg_row, start_column=1, end_row=avg_row, end_column=len(headers))
    
    wb.save(filepath)
    # Ensure file is fully written before serving
    wb.close()
    return filepath
